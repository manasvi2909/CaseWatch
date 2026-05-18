"""
excel_handler.py — Excel workbook operations for CaseWatch.

Handles:
  - Reading and parsing FIR/case records from .xlsx files
  - Multi-format date parsing
  - Row highlighting for overdue cases
  - Automatic backup before write operations
  - Graceful file-lock handling with retry logic
"""

import shutil
import time
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from app_logger import get_logger
from config import BASE_DIR, get_config
from models import Case

logger = get_logger("excel_handler")

# Backup directory
BACKUP_DIR = BASE_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# Column name mapping — case-insensitive, whitespace-tolerant
COLUMN_MAP = {
    "sno": ["sno", "s.no", "s no", "sr no", "serial", "sr.no", "sr. no"],
    "fsl_number": ["fsl number", "fsl no", "fslnumber", "fsl_number"],
    "fir_number": ["fir number", "fir no", "firnumber", "fir_number"],
    "police_station": ["police station", "ps", "policestation", "police_station"],
    "under_section": ["under section", "section", "undersection", "under_section"],
    "allotted_officer": [
        "allotted officer",
        "officer",
        "allottedofficer",
        "allotted_officer",
        "alloted officer",
    ],
    "date_receiving": [
        "date receiving",
        "date of receiving",
        "datereceiving",
        "date_receiving",
        "receiving date",
    ],
    "date_allotment": [
        "date allotment",
        "date of allotment",
        "dateallotment",
        "date_allotment",
        "allotment date",
    ],
    "date_reporting": [
        "date reporting",
        "date of reporting",
        "datereporting",
        "date_reporting",
        "reporting date",
    ],
}


def _normalize_column_name(name: str) -> str:
    """Normalize a column name for matching: lowercase, strip, collapse spaces."""
    return " ".join(str(name).lower().strip().split())


def _resolve_columns(df_columns: List[str]) -> dict:
    """
    Map DataFrame column names to internal field names.

    Returns:
        Dict mapping internal name -> actual DataFrame column name.

    Raises:
        ValueError: If required columns (date_allotment, date_reporting) are missing.
    """
    normalized = {_normalize_column_name(c): c for c in df_columns}
    mapping = {}

    for field_name, aliases in COLUMN_MAP.items():
        for alias in aliases:
            if alias in normalized:
                mapping[field_name] = normalized[alias]
                break

    # Validate required columns
    required = ["date_allotment", "date_reporting"]
    missing = [r for r in required if r not in mapping]
    if missing:
        raise ValueError(
            f"Required columns missing from Excel: {missing}. "
            f"Available columns: {list(df_columns)}"
        )

    return mapping


def _parse_date(value) -> Optional[date]:
    """
    Parse a date value from Excel, handling multiple formats.

    Supports:
      - datetime/date objects (from pandas/openpyxl)
      - String formats: DD/MM/YYYY, MM/DD/YYYY, YYYY-MM-DD, DD-MM-YYYY
      - NaT, None, empty strings → None
    """
    if value is None:
        return None

    # Check for pandas NaT FIRST — pd.NaT passes isinstance(date/datetime)
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # String parsing
    s = str(value).strip()
    if not s or s.lower() in ("nat", "none", "null", ""):
        return None

    # Try common date formats
    formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # Last resort: pandas parser
    try:
        parsed = pd.to_datetime(s, dayfirst=True)
        if not pd.isna(parsed):
            return parsed.date()
    except (ValueError, TypeError):
        pass

    return None


def scan_excel(file_path: str, sheet_name: Optional[str] = None) -> List[Case]:
    """
    Read and parse all FIR/case records from an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Name of the sheet to read. None = first sheet.

    Returns:
        List of Case objects with valid data.
        Rows with invalid/missing allotment dates are skipped (logged).
    """
    logger.info("Scanning Excel file: %s", file_path)
    path = Path(file_path)

    if not path.exists():
        logger.error("Excel file not found: %s", file_path)
        return []

    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name or 0,
            engine="openpyxl",
        )
    except Exception as e:
        logger.error("Failed to read Excel file: %s", e)
        return []

    if df.empty:
        logger.warning("Excel file is empty")
        return []

    # Resolve column mapping
    try:
        col_map = _resolve_columns(df.columns.tolist())
    except ValueError as e:
        logger.error("Column resolution failed: %s", e)
        return []

    logger.info("Column mapping resolved: %s", col_map)

    cases: List[Case] = []
    skipped = 0

    for idx, row in df.iterrows():
        # Excel row index: pandas index + 2 (header row + 0-based index)
        excel_row = int(idx) + 2

        try:
            # Skip entirely empty rows
            if row.isna().all():
                continue

            # Parse dates
            allotment = _parse_date(
                row.get(col_map.get("date_allotment"))
            ) if "date_allotment" in col_map else None

            reporting = _parse_date(
                row.get(col_map.get("date_reporting"))
            ) if "date_reporting" in col_map else None

            receiving = _parse_date(
                row.get(col_map.get("date_receiving"))
            ) if "date_receiving" in col_map else None

            # Skip rows without allotment date (can't determine overdue)
            if allotment is None:
                skipped += 1
                logger.debug(
                    "Row %d skipped: missing/invalid allotment date", excel_row
                )
                continue

            # Build case
            case = Case(
                row_index=excel_row,
                sno=_safe_int(row.get(col_map.get("sno"))),
                fsl_number=_safe_str(row.get(col_map.get("fsl_number"))),
                fir_number=_safe_str(row.get(col_map.get("fir_number"))),
                police_station=_safe_str(row.get(col_map.get("police_station"))),
                under_section=_safe_str(row.get(col_map.get("under_section"))),
                allotted_officer=_safe_str(row.get(col_map.get("allotted_officer"))),
                date_receiving=receiving,
                date_allotment=allotment,
                date_reporting=reporting,
            )
            cases.append(case)

        except Exception as e:
            skipped += 1
            logger.warning("Row %d skipped due to error: %s", excel_row, e)
            continue

    logger.info(
        "Scan complete: %d cases parsed, %d rows skipped", len(cases), skipped
    )
    return cases


def highlight_rows(
    file_path: str,
    overdue_rows: Set[int],
    reported_rows: Set[int],
    pending_rows: Set[int],
    overdue_color: str = "FFC7CE",
    reported_color: str = "C6EFCE",
    max_retries: int = 3,
    retry_delay: float = 2.0,
) -> bool:
    """
    Apply/remove highlighting on Excel rows.

    Creates a backup before any modifications.

    Args:
        file_path: Path to the .xlsx file.
        overdue_rows: Set of row indices to highlight red (1-based).
        reported_rows: Set of row indices to highlight green (1-based).
        pending_rows: Set of row indices to un-highlight (1-based).
        overdue_color: Hex color for overdue cases (without #).
        reported_color: Hex color for reported cases (without #).
        max_retries: Number of retry attempts if file is locked.
        retry_delay: Seconds between retries.

    Returns:
        True if highlighting was applied successfully.
    """
    if not overdue_rows and not reported_rows and not pending_rows:
        return True  # Nothing to do

    # Create backup first
    create_backup(file_path)

    fill_red = PatternFill(start_color=overdue_color, end_color=overdue_color, fill_type="solid")
    fill_green = PatternFill(start_color=reported_color, end_color=reported_color, fill_type="solid")
    fill_none = PatternFill(fill_type=None)

    for attempt in range(1, max_retries + 1):
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            max_col = ws.max_column

            # Apply overdue highlighting
            for row_idx in overdue_rows:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = fill_red

            # Apply reported highlighting
            for row_idx in reported_rows:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = fill_green

            # Remove highlighting for pending cases
            for row_idx in pending_rows:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = fill_none

            wb.save(file_path)
            wb.close()

            logger.info(
                "Highlighting updated: %d overdue (red), %d reported (green), %d pending (cleared)",
                len(overdue_rows),
                len(reported_rows),
                len(pending_rows),
            )
            return True

        except PermissionError:
            if attempt < max_retries:
                logger.warning(
                    "Excel file locked (attempt %d/%d), retrying in %.1fs...",
                    attempt,
                    max_retries,
                    retry_delay,
                )
                time.sleep(retry_delay)
            else:
                logger.error(
                    "Excel file locked after %d attempts. "
                    "Highlighting skipped (notifications will still fire).",
                    max_retries,
                )
                return False

        except Exception as e:
            logger.error("Failed to update highlighting: %s", e)
            return False

    return False


def create_backup(file_path: str) -> Optional[str]:
    """
    Create a timestamped backup of the Excel file.

    Maintains a maximum number of backups as configured.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Path to the backup file, or None if backup failed.
    """
    config = get_config()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    source = Path(file_path)
    if not source.exists():
        logger.warning("Cannot backup: source file does not exist: %s", file_path)
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{source.stem}_backup_{timestamp}{source.suffix}"
    backup_path = BACKUP_DIR / backup_name

    try:
        shutil.copy2(file_path, backup_path)
        logger.info("Backup created: %s", backup_path)
    except Exception as e:
        logger.error("Backup failed: %s", e)
        return None

    # Prune old backups
    _prune_backups(config.max_backups)

    return str(backup_path)


def _prune_backups(max_count: int) -> None:
    """Remove oldest backups beyond max_count."""
    if not BACKUP_DIR.exists():
        return

    backups = sorted(
        BACKUP_DIR.glob("*_backup_*.*"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    for old_backup in backups[max_count:]:
        try:
            old_backup.unlink()
            logger.debug("Pruned old backup: %s", old_backup.name)
        except OSError as e:
            logger.warning("Failed to prune backup %s: %s", old_backup.name, e)


def _safe_str(value) -> Optional[str]:
    """Safely convert a value to string, returning None for NaN/None."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    return s if s else None


def _safe_int(value) -> Optional[int]:
    """Safely convert a value to int, returning None for NaN/None."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None
