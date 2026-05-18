"""
generate_test_data.py — Generate a 100-row stress test Excel workbook for CaseWatch.
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment


def generate_test_excel(output_path: str = "test_data.xlsx") -> None:
    """Generate a sample 100-row stress test Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIR Records"

    # Headers
    headers = [
        "Sno",
        "FSL Number",
        "FIR Number",
        "Police Station",
        "Under Section",
        "Allotted Officer",
        "Date Receiving",
        "Date Allotment",
        "Date Reporting",
    ]

    header_font = Font(bold=True, size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    today = date.today()
    cases = []
    
    # 1. Overdue Cases (39 cases, rows 2-40)
    for i in range(1, 40):
        cases.append((
            i,
            f"FSL/2026/{100+i:03d}",
            f"{i}/26",
            "Sadar PS" if i % 2 == 0 else "Kotwali PS",
            "302 IPC" if i % 3 == 0 else "376 IPC",
            f"Officer {chr(65 + (i % 6))}",
            today - timedelta(days=25),
            today - timedelta(days=20),
            None
        ))

    # 2. Resolved Cases (30 cases, rows 41-70)
    for i in range(40, 70):
        cases.append((
            i,
            f"FSL/2026/{100+i:03d}",
            f"{i}/26",
            "Civil Lines PS" if i % 2 == 0 else "City PS",
            "420 IPC" if i % 3 == 0 else "379 IPC",
            f"Officer {chr(65 + (i % 6))}",
            today - timedelta(days=30),
            today - timedelta(days=25),
            today - timedelta(days=5)
        ))

    # 3. Pending but not overdue (10 cases, rows 71-80)
    for i in range(70, 80):
        cases.append((
            i,
            f"FSL/2026/{100+i:03d}",
            f"{i}/26",
            "Cantt PS",
            "323 IPC",
            f"Officer {chr(65 + (i % 6))}",
            today - timedelta(days=5),
            today - timedelta(days=3),
            None
        ))

    # 4. Future dates (5 cases, rows 81-85)
    for i in range(80, 85):
        cases.append((
            i,
            f"FSL/2026/{100+i:03d}",
            f"{i}/26",
            "Model Town PS",
            "498A IPC",
            "Officer Z",
            today + timedelta(days=5),
            today + timedelta(days=7),
            None
        ))

    # 5. Duplicates (5 cases, rows 86-90)
    # Re-adding exact copies of early overdue cases to verify duplicate scanning safety
    for i in range(85, 90):
        idx = i - 84 # duplicate of first few overdue cases
        cases.append((
            i,
            f"FSL/2026/{100+idx:03d}", # duplicate FSL
            f"{idx}/26",               # duplicate FIR
            "Sadar PS",
            "376 IPC",
            "Officer A",
            today - timedelta(days=25),
            today - timedelta(days=20),
            None
        ))

    # Write all valid / patterned cases to the sheet
    row = 2
    for case_data in cases:
        for col, value in enumerate(case_data, 1):
            cell = ws.cell(row=row, column=col)
            if value is not None:
                cell.value = value
                if isinstance(value, date):
                    cell.number_format = "DD/MM/YYYY"
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # 6. Blank Rows (3 rows: 91, 92, 93)
    # We will leave these entirely empty.
    row += 3

    # 7. Malformed Rows (7 rows: rows 94 to 100)
    malformed_cases = [
        # Row 94: Missing allotment date
        (91, "FSL/2026/901", "901/26", "Sadar PS", "302 IPC", "Officer X", today - timedelta(days=5), None, None),
        # Row 95: Text strings instead of dates in date allotment
        (92, "FSL/2026/902", "902/26", "Cantt PS", "376 IPC", "Officer Y", "invalid-date", "not-a-date", None),
        # Row 96: Non-numeric / missing FSL number (overdue, fallback to FIR)
        (93, None, "903/26", "City PS", "420 IPC", "Officer Z", today - timedelta(days=20), today - timedelta(days=15), None),
        # Row 97: String in Sno field
        ("SNO-MALFORMED", "FSL/2026/904", "904/26", "Kotwali PS", "302 IPC", "Officer A", today - timedelta(days=5), today - timedelta(days=4), None),
        # Row 98: Fully text columns
        (95, "FSL/TEXT/905", "905/26", "Sadar PS", "376 IPC", "Officer B", "not_a_date", "not_a_date", "not_a_date"),
        # Row 99: Date Allotment empty but has Date Reporting
        (96, "FSL/2026/906", "906/26", "Cantt PS", "307 IPC", "Officer C", today - timedelta(days=5), None, today - timedelta(days=1)),
        # Row 100: All columns empty except Sno and FSL number
        (97, "FSL/2026/907", None, None, None, None, None, None, None),
    ]

    for case_data in malformed_cases:
        for col, value in enumerate(case_data, 1):
            cell = ws.cell(row=row, column=col)
            if value is not None:
                cell.value = value
                if isinstance(value, date):
                    cell.number_format = "DD/MM/YYYY"
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = 18

    wb.save(output_path)
    print(f"Stress-test workbook generated successfully at: {output_path}")
    print(f"  - Total rows: {row - 1}")
    print(f"  - Overdue cases: 39 (rows 2-40)")
    print(f"  - Resolved cases: 30 (rows 41-70)")
    print(f"  - Pending (not overdue): 10 (rows 71-80)")
    print(f"  - Future date cases: 5 (rows 81-85)")
    # Note: 86-90 are duplicates of rows 2-6
    print(f"  - Duplicate cases: 5 (rows 86-90)")
    print(f"  - Blank rows: 3 (rows 91-93)")
    print(f"  - Malformed/Edge cases: 7 (rows 94-100)")


if __name__ == "__main__":
    generate_test_excel()
